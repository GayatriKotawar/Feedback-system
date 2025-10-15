import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

filepath = r"C:\Users\HP\Desktop\gayatri\Book1.xlsx"

# --- Setup Excel file with headers ---
def setup_excel():
    if not os.path.exists(filepath):
        try:
            workbook = Workbook()
            sheet = workbook.active
            headers = ["Timestamp"] + [f"Q{i+1}" for i in range(10)] + ["Summary"]
            sheet.append(headers)
            workbook.save(filepath)
        except PermissionError:
            messagebox.showerror("Permission Error", f"Cannot create file at {filepath}.\nClose the file if open or check permissions.")
            return False
    return True

responses = []

# --- Window 3: Thank you + Analysis ---
def show_results():
    happy = responses.count("😀")
    neutral = responses.count("😑")
    sad = responses.count("😔")

    # Determine overall sentiment
    if happy > neutral and happy > sad:
        summary = "Happy"
    elif neutral >= happy and neutral > sad:
        summary = "Neutral"
    else:
        summary = "Unhappy"

    # Save data to Excel
    try:
        workbook = load_workbook(filepath)
        sheet = workbook.active
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.append([timestamp] + responses + [summary])
        workbook.save(filepath)
    except PermissionError:
        messagebox.showerror("Permission Error", f"Cannot save to file {filepath}.\nMake sure it's closed and you have write permissions.")
        return

    # Show results window
    result_win = tk.Toplevel()
    result_win.title("Survey Results")
    result_win.configure(bg='white')
    result_win.geometry("450x400")
    result_win.resizable(False, False)

    tk.Label(result_win, text="Thank You for Submitting the Survey",
             font=("Arial", 14, "bold"), fg="blue", bg="white").pack(pady=20)

    tk.Label(result_win, text="Survey Analysis:", font=("Arial", 12, "bold"), fg="black", bg="white").pack(pady=10)

    tk.Label(result_win, text=f"😀 Happy: {happy}", font=("Arial", 12), bg="white").pack()
    tk.Label(result_win, text=f"😑 Neutral: {neutral}", font=("Arial", 12), bg="white").pack()
    tk.Label(result_win, text=f"😔 Sad: {sad}", font=("Arial", 12), bg="white").pack()

    tk.Label(result_win, text=f"\nOverall Sentiment: {summary}", font=("Arial", 12, "bold"), fg="black", bg="white").pack(pady=10)

    # Restart survey button
    def restart_survey():
        result_win.destroy()
        open_survey_window()

    tk.Button(result_win, text="Restart Survey", font=("Arial", 12, "bold"), bg="black", fg="white",
              command=restart_survey).pack(pady=15)

# --- Window 2: Survey ---
def open_survey_window():
    if not setup_excel():
        return  # Exit if Excel file can't be prepared

    survey_win = tk.Toplevel()
    survey_win.title("Workshop Feedback Survey")
    survey_win.configure(bg='white')
    survey_win.geometry("750x900")
    survey_win.resizable(False, False)

    questions = [
        "1. How was the overall workshop?",
        "2. How satisfied are you with the workshop content?",
        "3. How clear were the trainer’s explanations?",
        "4. How engaging was the trainer?",
        "5. How useful were the coding exercises?",
        "6. How well did the workshop meet your expectations?",
        "7. How comfortable did you feel asking questions during the session?",
        "8. How useful were the examples and real-life applications explained?",
        "9. How confident do you feel about applying Python after this workshop?",
        "10. How would you rate the clarity of concepts explained in the session?"
    ]

    responses.clear()
    emoji_buttons = []

    def handle_response(q_index, emoji):
        responses[q_index] = emoji
        # Reset button colors
        for btn in emoji_buttons[q_index]:
            btn.config(bg="white")
        # Highlight selected
        emoji_map = {"😀": 0, "😑": 1, "😔": 2}
        emoji_buttons[q_index][emoji_map[emoji]].config(bg="lightblue")

    for i, question in enumerate(questions):
        tk.Label(survey_win, text=question, font=("Arial", 11, "bold"), bg="white",
                 anchor="w", justify="left", wraplength=700).pack(padx=20, pady=(10, 2), anchor="w")

        frame = tk.Frame(survey_win, bg="white")
        frame.pack(pady=2, anchor="w", padx=40)

        responses.append("")

        b1 = tk.Button(frame, text="😀 Happy", font=("Arial", 11), bg="white",
                       command=lambda i=i: handle_response(i, "😀"))
        b2 = tk.Button(frame, text="😑 Neutral", font=("Arial", 11), bg="white",
                       command=lambda i=i: handle_response(i, "😑"))
        b3 = tk.Button(frame, text="😔 Sad", font=("Arial", 11), bg="white",
                       command=lambda i=i: handle_response(i, "😔"))

        b1.pack(side="left", padx=5)
        b2.pack(side="left", padx=5)
        b3.pack(side="left", padx=5)

        emoji_buttons.append([b1, b2, b3])

    def submit_survey():
        if "" in responses:
            messagebox.showwarning("Incomplete", "Please answer all questions before submitting.")
        else:
            survey_win.destroy()
            show_results()

    tk.Button(survey_win, text="Submit", font=("Arial", 13, "bold"),
              bg="blue", fg="white", command=submit_survey).pack(pady=30)

# --- Window 1: Main ---
def main_window():
    root = tk.Tk()
    root.title("Welcome to Python Workshop Survey")
    root.geometry("450x300")
    root.configure(bg="black")
    root.resizable(False, False)

    tk.Label(root, text="Python Workshop Feedback", font=("Arial", 18, "bold"),
             fg="white", bg="black").pack(pady=40)

    tk.Button(root, text="Start Survey", font=("Arial", 13, "bold"),
              bg="blue", fg="white", padx=20, pady=10,
              command=open_survey_window).pack()

    root.mainloop()

# --- Run the program ---
if __name__ == "__main__":
    main_window()
